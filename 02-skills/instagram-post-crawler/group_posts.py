#!/usr/bin/env python3
"""
Instagram 게시물 캡션 주제별 그룹화 — 크롤러 출력(posts.json)을 받아 주제(aspect)별로 묶는다.

쿠팡 group_reviews.py 의 그룹화 로직을 그대로 차용하되 인스타에 맞게 조정:
  - 텍스트 소스 = caption (리뷰 body 대신)
  - 리뷰엔 평점(rating)이 있어 긍/부정을 갈랐지만, 인스타 게시물엔 평점이 없다.
    → 대신 참여도(engagement = like + comment)로 상/중/하를 가른다(임의 감성 부여 금지).

결정적(외부 API 불필요) 산출 3종:
  1) 렉시콘 기반 주제 분류 — 주제→키워드 사전으로 캡션을 다중 라벨 분류.
     주제별 건수·평균 참여도·대표 스니펫. --topics 로 사전 교체 가능.
  2) 창발 키워드 — 사전에 없는 자주 나오는 단어/바이그램 상위 N(숨은 주제·해시태그 발견).
  3) 해시태그 집계 — 캡션 내 #태그 빈도.
"""
import argparse, collections, glob, json, os, re, subprocess, sys


def _reexec_into_python_with_openpyxl():
    """--xlsx 인데 현재 인터프리터에 openpyxl 이 없으면 있는 파이썬으로 재실행."""
    try:
        import openpyxl  # noqa
        return
    except Exception:
        pass
    if os.environ.get("_IGP_REEXEC"):
        return  # 무한 재실행 방지
    cands = []
    for pat in ("python3.13", "python3.12", "python3.11", "python3.10"):
        cands += glob.glob(f"/opt/homebrew/opt/python@*/bin/{pat}")
        cands += glob.glob(f"/usr/local/opt/python@*/bin/{pat}")
        cands += [f"/opt/homebrew/bin/{pat}", f"/usr/bin/{pat}"]
    for py in cands:
        if not os.path.exists(py):
            continue
        rc = subprocess.run([py, "-c", "import openpyxl"],
                            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL).returncode
        if rc == 0:
            os.environ["_IGP_REEXEC"] = "1"
            os.execv(py, [py, os.path.abspath(__file__), *sys.argv[1:]])
    sys.stderr.write("[fatal] openpyxl 이 있는 파이썬을 못 찾음. 먼저 설치: pip install -U openpyxl\n")
    sys.exit(2)

# 기본 렉시콘 — 인스타 브랜드/마케팅 게시물 범용. --topics custom.json 으로 교체/확장.
DEFAULT_TOPICS = {
    "제품·출시":   ["출시", "신제품", "런칭", "new", "출근", "제품", "라인업", "컬렉션", "입고"],
    "이벤트·경품": ["이벤트", "경품", "추첨", "당첨", "참여", "응모", "giveaway", "선물", "증정"],
    "할인·프로모": ["할인", "세일", "쿠폰", "특가", "프로모", "sale", "혜택", "기간한정", "오픈"],
    "후기·리뷰":   ["후기", "리뷰", "사용해", "써봤", "만족", "추천", "재구매", "review"],
    "정보·팁":     ["방법", "팁", "가이드", "how", "노하우", "정보", "알아", "튜토리얼", "꿀팁"],
    "브랜드·소식": ["소식", "공지", "안내", "오픈", "매장", "입점", "파트너", "협업", "콜라보"],
    "일상·비하인드": ["일상", "비하인드", "daily", "오늘", "스탭", "현장", "촬영", "behind"],
    "고객·커뮤니티": ["고객", "여러분", "팔로우", "댓글", "소통", "감사", "함께", "community"],
    "채용·팀":     ["채용", "합류", "팀원", "모집", "recruit", "커리어", "인턴", "지원"],
    "성과·수상":   ["성과", "수상", "선정", "달성", "1위", "award", "돌파", "기록", "누적"],
}

_WORD = re.compile(r"[가-힣A-Za-z0-9]{2,}")
_HASHTAG = re.compile(r"#([0-9A-Za-z_가-힣]+)")
_STOP = set("그리고 그래서 하지만 정말 너무 진짜 조금 아주 매우 그냥 이거 저거 이것 그것 "
            "해서 하고 있어요 좋아요 같아요 합니다 어요 네요 에서 으로 지금 오늘 우리 여기 "
            "instagram insta 인스타 그램 https com www".split())


def engagement(r):
    like = r.get("like_count") or 0
    com = r.get("comment_count") or 0
    try:
        return int(like) + int(com)
    except (ValueError, TypeError):
        return 0


def snippet(r, n=90):
    b = (r.get("caption") or "").strip().replace("\n", " ")
    return (b[:n] + "…") if len(b) > n else b


def classify(posts, topics):
    groups = {t: {"posts": [], "eng": []} for t in topics}
    unmatched = []
    for r in posts:
        text = (r.get("caption") or "")
        hit = False
        for t, kws in topics.items():
            if any(kw.lower() in text.lower() for kw in kws):
                groups[t]["posts"].append(r)
                groups[t]["eng"].append(engagement(r))
                hit = True
        if not hit:
            unmatched.append(r)
    return groups, unmatched


def emergent_keywords(posts, topics, top=25):
    known = set()
    for kws in topics.values():
        known.update(k.lower() for k in kws)
    uni = collections.Counter()
    bi = collections.Counter()
    for r in posts:
        cap = r.get("caption") or ""
        # 해시태그·멘션 토큰 제거한 순수 단어
        clean = re.sub(r"[#@][0-9A-Za-z_가-힣]+", " ", cap)
        toks = [w for w in _WORD.findall(clean) if w.lower() not in known and w not in _STOP]
        for w in toks:
            uni[w] += 1
        for a, b in zip(toks, toks[1:]):
            bi[f"{a} {b}"] += 1
    return uni.most_common(top), bi.most_common(top)


def hashtag_agg(posts, top=40):
    c = collections.Counter()
    for r in posts:
        for m in _HASHTAG.findall(r.get("caption") or ""):
            c[m.lower()] += 1
    return c.most_common(top)


def _safe_sheet(name):
    for ch in "[]:*?/\\":
        name = name.replace(ch, " ")
    return name[:31]


def write_xlsx(path, posts, groups, out_json, hashtags):
    """주제별 탭 엑셀 — 요약 시트 + 주제마다 시트 1개(매칭 게시물 전체) + 해시태그 시트."""
    import openpyxl
    from openpyxl.styles import Font
    wb = openpyxl.Workbook()
    # 요약 시트
    ws = wb.active
    ws.title = "요약"
    ws.append(["주제", "게시물수", "평균참여도", "최대참여도"])
    for c in ws[1]:
        c.font = Font(bold=True)
    ranked = sorted(out_json["topics"].items(), key=lambda x: x[1]["count"], reverse=True)
    for t, d in ranked:
        ws.append([t, d["count"], d["avg_engagement"], d["max_engagement"]])
    ws.freeze_panes = "A2"
    # 주제별 시트 (게시물 많은 순)
    cols = ["shortcode", "url", "timestamp", "like_count", "comment_count",
            "is_video", "owner_username", "caption"]
    for t, _ in ranked:
        rows = groups[t]["posts"]
        if not rows:
            continue
        s = wb.create_sheet(_safe_sheet(t))
        s.append(cols)
        for c in s[1]:
            c.font = Font(bold=True)
        s.freeze_panes = "A2"
        for r in sorted(rows, key=engagement, reverse=True):
            s.append([r.get(k, "") for k in cols])
        s.column_dimensions["B"].width = 42
        s.column_dimensions["H"].width = 80
    # 해시태그 시트
    if hashtags:
        hs = wb.create_sheet("해시태그")
        hs.append(["해시태그", "빈도"])
        for c in hs[1]:
            c.font = Font(bold=True)
        for tag, n in hashtags:
            hs.append([f"#{tag}", n])
    wb.save(path)


def main():
    ap = argparse.ArgumentParser(description="Instagram 게시물 캡션 주제별 그룹화")
    ap.add_argument("posts_json", help="크롤러 출력 JSON (posts.json)")
    ap.add_argument("--topics", help="주제 사전 JSON (없으면 기본 렉시콘)")
    ap.add_argument("--out", default="posts_grouped", help="출력 파일 접두어(.json/.md/.xlsx)")
    ap.add_argument("--examples", type=int, default=3, help="주제별 대표 스니펫 수")
    ap.add_argument("--xlsx", action="store_true", help="주제별 탭 엑셀도 출력(openpyxl 필요)")
    args = ap.parse_args()

    if args.xlsx:
        _reexec_into_python_with_openpyxl()

    posts = json.load(open(args.posts_json, encoding="utf-8"))
    topics = json.load(open(args.topics, encoding="utf-8")) if args.topics else DEFAULT_TOPICS

    groups, unmatched = classify(posts, topics)
    uni, bi = emergent_keywords(posts, topics)
    hashtags = hashtag_agg(posts)

    # ── JSON 산출 ──
    out_json = {"total": len(posts), "unmatched": len(unmatched), "topics": {}}
    for t, g in groups.items():
        cnt = len(g["posts"])
        avg = round(sum(g["eng"]) / len(g["eng"]), 1) if g["eng"] else None
        out_json["topics"][t] = {
            "count": cnt,
            "avg_engagement": avg,
            "max_engagement": max(g["eng"]) if g["eng"] else None,
            "sample": [snippet(r) for r in
                       sorted(g["posts"], key=engagement, reverse=True)[:args.examples]],
        }
    out_json["emergent_unigrams"] = uni
    out_json["emergent_bigrams"] = bi
    out_json["hashtags"] = hashtags
    json.dump(out_json, open(args.out + ".json", "w", encoding="utf-8"),
              ensure_ascii=False, indent=2)

    # ── Markdown 리포트 ──
    L = [f"# Instagram 게시물 주제별 그룹화 — 총 {len(posts)}건", ""]
    ranked = sorted(out_json["topics"].items(), key=lambda x: x[1]["count"], reverse=True)
    L.append("## 주제별 요약 (게시물 많은 순)")
    L.append("")
    L.append("| 주제 | 게시물 | 평균 참여도 | 최대 참여도 |")
    L.append("|---|---:|---:|---:|")
    for t, d in ranked:
        L.append(f"| {t} | {d['count']} | "
                 f"{d['avg_engagement'] if d['avg_engagement'] is not None else '-'} | "
                 f"{d['max_engagement'] if d['max_engagement'] is not None else '-'} |")
    L.append("")
    for t, d in ranked:
        if d["count"] == 0:
            continue
        L.append(f"### {t} — {d['count']}건 (평균 참여도 {d['avg_engagement']})")
        for s in d["sample"]:
            L.append(f"- {s}")
        L.append("")
    if hashtags:
        L.append("## 해시태그 빈도 (상위)")
        L.append("- " + ", ".join(f"#{t}({n})" for t, n in hashtags[:20]))
        L.append("")
    L.append("## 창발 키워드 (사전 밖 자주 나온 표현)")
    L.append("- 단어: " + ", ".join(f"{w}({n})" for w, n in uni[:15]))
    L.append("- 연어(바이그램): " + ", ".join(f"{w}({n})" for w, n in bi[:15]))
    if unmatched:
        L.append("")
        L.append(f"> 미분류 {len(unmatched)}건 — 기본 사전에 안 걸림. 창발 키워드를 --topics 사전에 추가하면 흡수.")
    open(args.out + ".md", "w", encoding="utf-8").write("\n".join(L))

    if args.xlsx:
        write_xlsx(args.out + ".xlsx", posts, groups, out_json, hashtags)

    print(json.dumps({"total": len(posts),
                      "topics_ranked": [[t, d["count"]] for t, d in ranked],
                      "unmatched": len(unmatched),
                      "out_md": args.out + ".md", "out_json": args.out + ".json"},
                     ensure_ascii=False))


if __name__ == "__main__":
    main()
