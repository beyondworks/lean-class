#!/usr/bin/env python3
"""
Coupang 리뷰 주제별 그룹화 — 크롤러 출력(JSON)을 받아 주제(aspect)별로 묶고 리포트 생성.

3가지를 결정적(외부 API 불필요)으로 산출:
  1) 렉시콘 기반 주제 분류 — 주제→키워드 사전으로 리뷰를 다중 라벨 분류.
     주제별 건수·평균평점·긍/부정 분리·대표 스니펫. --topics 로 사전 교체 가능.
  2) 창발 키워드 — 사전에 없는 자주 나오는 단어/바이그램 상위 N (숨은 주제 발견용).
  3) 설문 집계 — 쿠팡 구조화 설문(향/효과/자극/보습/흡수 등) 항목별 응답 분포.

감성은 평점으로 가른다: rating>=4 긍정, <=2 부정, 3 중립.
LLM 이 이 위에 더 풍부한 의미 클러스터링을 얹을 수 있게 JSON 도 함께 출력.
"""
import argparse, collections, json, re, sys

# 기본 렉시콘 — 쿠팡 리뷰 범용 + 화장품 결. --topics custom.json 으로 교체/확장.
DEFAULT_TOPICS = {
    "효과·성능":   ["효과", "효능", "좋아졌", "개선", "달라", "변화", "확실히", "느껴", "성능"],
    "가격·가성비": ["가격", "가성비", "저렴", "싸", "비싸", "할인", "혜자", "값", "돈값", "가심비"],
    "배송":        ["배송", "도착", "빨리", "빠르", "로켓", "포장", "택배", "다음날"],
    "포장·용기":   ["용기", "포장", "패키지", "뚜껑", "펌프", "새", "파손", "누출", "깨졌"],
    "향·냄새":     ["향", "냄새", "향기", "무향", "은은", "인공", "역하", "좋은 냄새"],
    "사용감·편의": ["사용감", "바르", "발림", "촉촉", "산뜻", "끈적", "무겁", "가볍", "흡수"],
    "품질·내구성": ["품질", "튼튼", "내구", "고장", "불량", "하자", "마감", "오래"],
    "재구매·추천": ["재구매", "또 살", "추천", "만족", "재구입", "믿고", "선물", "강추"],
    "성분·안전":   ["성분", "자극", "트러블", "민감", "순하", "무자극", "저자극", "알러지", "따가"],
    "보습·수분":   ["보습", "수분", "촉촉", "건조", "속건조", "당김", "촉촉함", "쫀쫀"],
}

_WORD = re.compile(r"[가-힣A-Za-z0-9]{2,}")
_STOP = set("그리고 그래서 하지만 정말 너무 진짜 조금 아주 매우 그냥 제품 상품 리뷰 구매 사용 "
            "이거 저거 이것 그것 해서 하고 있어요 좋아요 같아요 합니다 어요 네요 에서 으로 "
            "이니스프리 쿠팡 체험단 이벤트 제공받아 작성 무료".split())


def sentiment(r):
    try:
        v = int(r.get("rating") or 0)
    except ValueError:
        v = 0
    return "pos" if v >= 4 else ("neg" if 1 <= v <= 2 else "neu")


def snippet(r, n=90):
    b = (r.get("body") or "").strip().replace("\n", " ")
    return (b[:n] + "…") if len(b) > n else b


def classify(reviews, topics):
    groups = {t: {"pos": [], "neg": [], "neu": [], "ratings": []} for t in topics}
    unmatched = []
    for r in reviews:
        text = (r.get("headline", "") + " " + r.get("body", "") + " " + r.get("survey", ""))
        hit = False
        for t, kws in topics.items():
            if any(kw in text for kw in kws):
                s = sentiment(r)
                groups[t][s].append(r)
                try:
                    groups[t]["ratings"].append(int(r.get("rating") or 0))
                except ValueError:
                    pass
                hit = True
        if not hit:
            unmatched.append(r)
    return groups, unmatched


def emergent_keywords(reviews, topics, top=25):
    known = set()
    for kws in topics.values():
        known.update(kws)
    uni = collections.Counter()
    bi = collections.Counter()
    for r in reviews:
        toks = [w for w in _WORD.findall((r.get("body") or "")) if w not in _STOP]
        for w in toks:
            if w not in known:
                uni[w] += 1
        for a, b in zip(toks, toks[1:]):
            bi[f"{a} {b}"] += 1
    return uni.most_common(top), bi.most_common(top)


def survey_agg(reviews):
    # survey: "향 만족도 아주만족해요 | 사용 효과 아주만족해요 | ..." → 항목별 응답 분포
    agg = collections.defaultdict(collections.Counter)
    for r in reviews:
        for cell in (r.get("survey") or "").split("|"):
            cell = cell.strip()
            if not cell:
                continue
            # 마지막 토큰 뭉치가 응답(…해요/…없어요), 앞이 항목명
            m = re.match(r"(.+?)\s+(\S+요|\S+음|\S+됨|\S+완화\S*|\S+흡수\S*)$", cell)
            if m:
                agg[m.group(1).strip()][m.group(2).strip()] += 1
    return agg


def _safe_sheet(name):
    for ch in "[]:*?/\\":
        name = name.replace(ch, " ")
    return name[:31]


def write_xlsx(path, reviews, groups, topics, out_json):
    """주제별 탭으로 나뉜 엑셀 — 요약 시트 + 주제마다 시트 1개(매칭 리뷰 전체)."""
    import openpyxl
    from openpyxl.styles import Font
    wb = openpyxl.Workbook()
    # 요약 시트
    ws = wb.active
    ws.title = "요약"
    ws.append(["주제", "언급", "평균★", "긍정", "부정", "중립"])
    for c in ws[1]:
        c.font = Font(bold=True)
    ranked = sorted(out_json["topics"].items(), key=lambda x: x[1]["count"], reverse=True)
    for t, d in ranked:
        ws.append([t, d["count"], d["avg_rating"], d["pos"], d["neg"], d["neu"]])
    ws.freeze_panes = "A2"
    # 주제별 시트 (언급 많은 순)
    cols = ["rating", "date", "user", "headline", "body", "survey", "helpful"]
    for t, _ in ranked:
        g = groups[t]
        rows = g["pos"] + g["neu"] + g["neg"]
        if not rows:
            continue
        s = wb.create_sheet(_safe_sheet(t))
        s.append(cols)
        for c in s[1]:
            c.font = Font(bold=True)
        s.freeze_panes = "A2"
        for r in sorted(rows, key=lambda x: x.get("rating", ""), reverse=True):
            s.append([r.get(k, "") for k in cols])
        # 본문 열 넓게
        s.column_dimensions["E"].width = 80
        s.column_dimensions["F"].width = 40
    wb.save(path)


def main():
    ap = argparse.ArgumentParser(description="Coupang 리뷰 주제별 그룹화")
    ap.add_argument("reviews_json", help="크롤러 출력 JSON")
    ap.add_argument("--topics", help="주제 사전 JSON (없으면 기본 렉시콘)")
    ap.add_argument("--out", default="coupang_reviews_grouped", help="출력 파일 접두어(.json/.md/.xlsx)")
    ap.add_argument("--examples", type=int, default=3, help="주제별 대표 스니펫 수")
    ap.add_argument("--xlsx", action="store_true", help="주제별 탭 엑셀도 출력(openpyxl 필요)")
    args = ap.parse_args()

    reviews = json.load(open(args.reviews_json, encoding="utf-8"))
    topics = json.load(open(args.topics, encoding="utf-8")) if args.topics else DEFAULT_TOPICS

    groups, unmatched = classify(reviews, topics)
    uni, bi = emergent_keywords(reviews, topics)
    survey = survey_agg(reviews)

    # ── JSON 산출 ──
    out_json = {"total": len(reviews), "unmatched": len(unmatched), "topics": {}}
    for t, g in groups.items():
        cnt = len(g["pos"]) + len(g["neg"]) + len(g["neu"])
        avg = round(sum(g["ratings"]) / len(g["ratings"]), 2) if g["ratings"] else None
        out_json["topics"][t] = {
            "count": cnt, "avg_rating": avg,
            "pos": len(g["pos"]), "neg": len(g["neg"]), "neu": len(g["neu"]),
            "sample_pos": [snippet(r) for r in g["pos"][:args.examples]],
            "sample_neg": [snippet(r) for r in g["neg"][:args.examples]],
        }
    out_json["emergent_unigrams"] = uni
    out_json["emergent_bigrams"] = bi
    out_json["survey"] = {k: dict(v) for k, v in survey.items()}
    json.dump(out_json, open(args.out + ".json", "w", encoding="utf-8"),
              ensure_ascii=False, indent=2)

    # ── Markdown 리포트 ──
    L = [f"# 쿠팡 리뷰 주제별 그룹화 — 총 {len(reviews)}건", ""]
    ranked = sorted(out_json["topics"].items(), key=lambda x: x[1]["count"], reverse=True)
    L.append("## 주제별 요약 (언급 많은 순)")
    L.append("")
    L.append("| 주제 | 언급 | 평균★ | 긍정 | 부정 |")
    L.append("|---|---:|---:|---:|---:|")
    for t, d in ranked:
        L.append(f"| {t} | {d['count']} | {d['avg_rating'] if d['avg_rating'] is not None else '-'} | {d['pos']} | {d['neg']} |")
    L.append("")
    for t, d in ranked:
        if d["count"] == 0:
            continue
        L.append(f"### {t} — {d['count']}건 (평균 {d['avg_rating']}★, 긍정 {d['pos']} / 부정 {d['neg']})")
        if d["sample_pos"]:
            L.append("- 긍정 예시:")
            L += [f"  - {s}" for s in d["sample_pos"]]
        if d["sample_neg"]:
            L.append("- 부정 예시:")
            L += [f"  - {s}" for s in d["sample_neg"]]
        L.append("")
    if survey:
        L.append("## 쿠팡 설문 항목 집계")
        L.append("")
        for k, v in survey.items():
            dist = ", ".join(f"{resp} {n}" for resp, n in v.most_common())
            L.append(f"- **{k}**: {dist}")
        L.append("")
    L.append("## 창발 키워드 (사전 밖 자주 나온 표현)")
    L.append("- 단어: " + ", ".join(f"{w}({n})" for w, n in uni[:15]))
    L.append("- 연어(바이그램): " + ", ".join(f"{w}({n})" for w, n in bi[:15]))
    if unmatched:
        L.append("")
        L.append(f"> 미분류 {len(unmatched)}건 — 기본 사전에 안 걸림. 창발 키워드를 --topics 사전에 추가하면 흡수.")
    open(args.out + ".md", "w", encoding="utf-8").write("\n".join(L))

    if args.xlsx:
        write_xlsx(args.out + ".xlsx", reviews, groups, topics, out_json)

    print(json.dumps({"total": len(reviews),
                      "topics_ranked": [[t, d["count"]] for t, d in ranked],
                      "unmatched": len(unmatched),
                      "out_md": args.out + ".md", "out_json": args.out + ".json"},
                     ensure_ascii=False))


if __name__ == "__main__":
    main()
