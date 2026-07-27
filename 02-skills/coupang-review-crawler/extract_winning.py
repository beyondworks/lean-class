#!/usr/bin/env python3
"""
위닝 소재 추출 — 로우 리뷰에서 '리뷰어의 결핍(고객의 문제·컴플렉스)'을 결핍-우선으로
뽑아 빈도집계하고, 반복되는 결핍을 소구포인트·후킹 카피로 승격한다.

group_reviews.py 와의 차이(관점 반전):
  - group_reviews = 제품 강점/주제 중심(무엇이 좋은가).
  - extract_winning = 고객의 결핍 중심(무엇이 아팠는가 → 제품이 그걸 어떻게 푸는가).
    광고 소재는 제품 자랑이 아니라 '고객의 통증'에서 후킹되므로, 결핍이 위닝의 씨앗이다.

정밀도가 생명 (안 그러면 전체가 오도됨). 두 종류의 결핍을 다르게 다룬다:
  - TRAIT  = 고객이 원래 가진 조건·컴플렉스(직모·짧은 속눈썹·숱없음·눈두덩 먹힘·건성…).
             언급 자체가 결핍이다(문장이 긍정이어도 "나는 직모인데"는 결핍 컨텍스트).
             → 위닝 소재의 핵심. "직모라 포기했는데 이건…" 이 바로 이 유형.
  - OUTCOME= 증상(번짐·뭉침·처짐·금방지워짐). 제품이 유발도 해결도 할 수 있어 **부정어에
             민감**하다. "번짐 없음"은 강점(제품이 해결), "번져서 창피"는 결핍.
             → 부정어(없/안/않/덜/전혀) 게이트를 통과한 것만 결핍으로 카운트.
결정적·재현 가능. LLM 이 이 위에 최종 카피를 다듬는다(스크립트는 결핍을 정밀 격리하는 역할).
"""
import argparse, collections, json, re, sys

# 결핍 사전 — {상황: {"kw":[표현], "kind":"trait|outcome"}}. --deficits custom.json 으로 교체.
# 각 키 = 고객이 공통으로 겪는 '상황(문제)'. 비슷한 표현은 한 상황으로 묶어 건수로 센다
# (예: "힘없이 내려가는"·"처지는"·"뻐신 직모"는 모두 '속눈썹이 힘없이 처짐' 한 그룹).
# 뷰티(외모 컴플렉스) 기본값. 상품군 다르면 커스텀 사전을 넘겨라.
DEFAULT_DEFICITS = {
    "속눈썹이 짧고 숱이 적음":     {"kind": "trait",   "kw": ["짧은 속눈썹", "속눈썹 짧", "속눈썹이 짧", "짧고", "숱 적", "숱적", "숱 없", "숱없", "숱이 없", "빈약", "몇 가닥", "티도 안", "티가 안"]},
    # 주의: '쳐지/쳐져'는 '뭉쳐져/뭉쳐지'와, '아래로/내려가'는 '아래로 번짐'과 부분문자열 충돌 → 제외.
    #       '처져/처지는' 등 ㅊ 형태 + 힘없/직모/뻐신 으로 상황을 잡는다(런온 리뷰도 이들로 충분히 커버).
    "속눈썹이 힘없이 처지고 직모": {"kind": "trait",   "kw": ["힘없", "힘 없", "힘이 없", "기운 없", "처지는", "처지고", "처진", "처져", "축 처", "직모", "일자 속눈썹", "안 올라가", "안올라가", "뻐신", "뻣뻣", "뻐센"]},
    "두꺼운 솔이 눈두덩·눈가에 묻음": {"kind": "trait", "kw": ["눈두덩", "눈덩이", "눈가에 묻", "면봉으로 닦", "묻히곤", "다 묻어", "묻어서 곤란", "솔이 크", "솔이 두"]},
    "유분 많아 번짐·팬더 됨":     {"kind": "outcome", "kw": ["번져", "번지", "번짐", "판다", "팬더", "눈밑 검", "눈 밑 검", "눈밑이 다", "기름기", "유분", "다크서클처럼"]},
    "뭉침·떡짐":                 {"kind": "outcome", "kw": ["뭉쳐", "뭉침", "떡져", "떡짐", "파리 다리", "파리다리", "덕지"]},
    "금방 지워짐·수정화장 잦음":  {"kind": "outcome", "kw": ["금방 지워", "몇 시간 만", "무너져", "수정 화장", "지속력 없", "오래 못 가", "쉽게 지워", "유지 안"]},
    "속눈썹 빠짐·손상":          {"kind": "trait",   "kw": ["많이 빠", "속눈썹이 빠", "빠져서", "빠지고", "손상", "연장 부작용", "약해져", "가늘어"]},
    "눈매가 답답·작아 보임":     {"kind": "trait",   "kw": ["무쌍", "속쌍", "졸려 보", "작은 눈", "작아 보", "답답해 보", "또렷하지 않", "흐릿"]},
    "건조·각질":                {"kind": "trait",   "kw": ["건조", "각질", "당김", "속건조", "푸석"]},
    "트러블·자극·민감":         {"kind": "trait",   "kw": ["트러블", "자극", "따가", "예민", "민감", "뒤집어", "붉어", "간지러"]},
    "마스카라 자체를 포기함":    {"kind": "trait",   "kw": ["포기하고 살", "포기하고 지", "마스카라 포기", "화장 포기", "안 바르다", "안바르다", "안 쓰다가", "숨기고", "가리고"]},
}

# 체험단·협찬 리뷰 — 광고 편향이라 VOC 에서 제외(기본). --keep-promo 로 유지 가능.
PROMO_MARKERS = ["체험단", "제공받아", "제공 받아", "제공 받은", "제공받은", "무상", "무료로 제공",
                 "무료 제공", "무료로 받", "협찬", "서포터즈", "소정의", "원고료", "대가를 받",
                 "리뷰를 작성", "작성한 후기", "작성된 후기", "이벤트 당첨", "이벤트에 당첨", "업체로부터"]

# 부정/해결 토큰 — OUTCOME 결핍 문장에 이게 있으면 '제품이 해결한 강점'이라 결핍서 제외.
# 한국어 부정은 문장 끝 임의 거리에 오므로(뭉침이 하나도 '없') 문장 전체를 본다.
_NEG = ["없", "않", "안 ", "덜 ", "덜하", "전혀", "일절", "제로", "강한", "강해", "걱정 없"]
# 통증 프레임 — OUTCOME 증상이 '진짜 불만'임을 확증하는 감정·지속 마커(같은 문장 동반 필수).
PAIN_FRAME = ["고민", "컴플렉스", "콤플렉스", "스트레스", "창피", "수치", "부끄", "포기",
              "매번", "늘 ", "항상", "평소", "원래", "예전", "때문", "곤란", "속상", "고생",
              "ㅠ", "ㅜ", "자주 했", "자주했", "신경", "걱정", "짜증", "실패", "골치"]
# 반전(해결) 마커 — 같은 문장 안 결핍→해결 아크 감지용.
RESOLVE_MARKERS = ["이건", "이거는", "이제품", "이제", "드디어", "처음", "확실히", "신세계",
                   "미쳤", "인생", "해결", "살렸", "살아", "또렷", "커 보", "선명", "신기",
                   "덕분", "이것만", "바르자마자", "쫙", "바짝", "올라가", "길어져", "정착"]
_WORD = re.compile(r"[가-힣]{2,}")
_SENT = re.compile(r"[^.!?~\n。]+[.!?~\n。]?")
_STOP = set("그리고 그래서 하지만 정말 너무 진짜 조금 아주 매우 그냥 제품 상품 리뷰 구매 사용 "
            "이거 저거 이것 그것 해서 하고 있어요 좋아요 같아요 합니다 에서 으로 마스카라 "
            "쿠팡 체험단 이벤트 제공받아 작성 후기 배송 가격 정말로 그램".split())


def _rating(r):
    try:
        return int(r.get("rating") or 0)
    except (ValueError, TypeError):
        return 0


def sentences(text):
    return [s.strip() for s in _SENT.findall(text or "") if s.strip()]


def _clean_clause(sent, kw):
    """상황 키워드 주변의 깔끔한 클로즈만 뽑는다(런온 리뷰에서 문구가 파묻히지 않게).
    광고 소재로 바로 쓰도록 짧고 문맥 있는 원문 조각을 돌려준다."""
    s = sent.strip()
    if len(s) <= 70:
        return s
    i = s.find(kw)
    if i < 0:
        return s[:70].strip() + "…"
    left = 0
    for m in re.finditer(r"[,.!?~。]", s[:i]):
        left = m.end()
    right = len(s)
    m = re.search(r"[,.!?~。]", s[i + len(kw):])
    if m:
        right = i + len(kw) + m.start() + 1
    clause = s[left:right].strip()
    if 8 <= len(clause) <= 80:
        return clause
    a, b = max(0, i - 32), min(len(s), i + len(kw) + 40)
    return ("…" if a > 0 else "") + s[a:b].strip() + ("…" if b < len(s) else "")


def emotion_score(text):
    t = text or ""
    return sum(t.count(m) for m in ("ㅠ", "ㅜ", "!", "포기", "창피", "수치", "미쳤",
                                    "스트레스", "고민", "곤란", "속상"))


def _negated(sent, kw):
    """kw 위치 이후 문장에 부정/강점 토큰이 있으면 True (한국어 부정은 문장 끝에 옴)."""
    i = sent.find(kw)
    if i < 0:
        return False
    after = sent[i + len(kw):]
    return any(n.strip() in after for n in _NEG) or "지 않" in after


def deficit_hit(text, theme, spec):
    """이 리뷰가 해당 결핍을 가지면 대표 결핍 문장을 반환, 아니면 None.
    TRAIT   = 고객의 조건 언급만으로 결핍(견고, 위닝 핵심).
    OUTCOME = 증상 → 같은 문장에 통증 프레임 있고 부정어 없을 때만(정밀도 우선)."""
    is_outcome = spec["kind"] == "outcome"
    for s in sentences(text):
        for kw in spec["kw"]:
            if kw not in s:
                continue
            if is_outcome:
                if _negated(s, kw):
                    continue  # '뭉침 없음' = 강점
                if not any(p in s for p in PAIN_FRAME):
                    continue  # 통증 프레임 없으면 진짜 불만인지 불확실 → 제외
            return _clean_clause(s, kw)  # 상황 문구 중심의 깔끔한 원문 조각
    return None


def resolve_sentence(text, deficit_sent):
    """같은 문장 안 반전(해결) 아크만 — 결핍 문장이 해결까지 담으면 그 문장, 아니면 빈값.
    문장을 넘는 결핍↔해결 페어링은 신뢰 불가라 하지 않는다(오도 방지)."""
    if deficit_sent and any(m in deficit_sent for m in RESOLVE_MARKERS):
        return deficit_sent  # 한 문장에 결핍+해결 아크 ("눈두덩이 묻히곤 했는데 이건…또렷")
    return ""


def classify(reviews, deficits):
    groups = {t: {"reviews": [], "quotes": []} for t in deficits}
    pain_only = []
    PAIN = ["고민", "컴플렉스", "콤플렉스", "스트레스", "포기", "창피", "수치", "부끄",
            "자신 없", "신경 쓰", "답답", "고생", "불편", "효과 없", "안 맞", "매번", "늘 ", "항상"]
    for r in reviews:
        text = (r.get("headline", "") + " " + r.get("body", "")).strip()
        hit = False
        for t, spec in deficits.items():
            ds = deficit_hit(text, t, spec)
            if ds:
                groups[t]["reviews"].append(r)
                groups[t]["quotes"].append({
                    "user": r.get("user", ""), "rating": _rating(r),
                    "deficit": ds, "resolve": resolve_sentence(text, ds),
                    "emotion": emotion_score(text), "kind": spec["kind"],
                })
                hit = True
        if not hit and any(p in text for p in PAIN):
            pain_only.append(r)
    return groups, pain_only


def emergent(pain_reviews, deficits, top=20):
    known = set(w for spec in deficits.values() for kw in spec["kw"] for w in kw.split())
    uni, bi = collections.Counter(), collections.Counter()
    for r in pain_reviews:
        toks = [w for w in _WORD.findall(r.get("body") or "") if w not in _STOP and w not in known]
        for w in toks:
            uni[w] += 1
        for a, b in zip(toks, toks[1:]):
            bi[f"{a} {b}"] += 1
    return uni.most_common(top), bi.most_common(top)


def main():
    ap = argparse.ArgumentParser(description="쿠팡 리뷰 위닝 소재 추출(결핍-우선)")
    ap.add_argument("reviews_json", help="크롤러 출력 JSON")
    ap.add_argument("--deficits", help="결핍 주제 사전 JSON. 없으면 뷰티 기본값")
    ap.add_argument("--out", default="coupang_winning", help="출력 접두어(.md/.json[/.xlsx])")
    ap.add_argument("--top", type=int, default=6, help="소구/후킹 도출 상위 결핍 수")
    ap.add_argument("--quotes", type=int, default=20, help="상황당 대표 원문 인용 수(최대, 기본 20)")
    ap.add_argument("--xlsx", action="store_true", help="결핍별 탭 엑셀도 출력(openpyxl 필요)")
    ap.add_argument("--keep-promo", action="store_true", help="체험단·협찬 리뷰 유지(기본은 제외)")
    args = ap.parse_args()

    raw = json.load(open(args.reviews_json, encoding="utf-8"))
    deficits = json.load(open(args.deficits, encoding="utf-8")) if args.deficits else DEFAULT_DEFICITS

    # 본문 없는 별점-only + (기본)체험단·협찬 제거 → 진짜 고객 VOC 만 분석
    n_raw = len(raw)
    reviews = [r for r in raw if (r.get("body") or "").strip()]
    n_empty = n_raw - len(reviews)
    n_promo = 0
    if not args.keep_promo:
        kept = [r for r in reviews if not any(m in (r.get("body") or "") for m in PROMO_MARKERS)]
        n_promo = len(reviews) - len(kept)
        reviews = kept
    filt = {"raw": n_raw, "empty_body": n_empty, "promo": n_promo, "analyzed": len(reviews)}

    groups, pain_only = classify(reviews, deficits)
    uni, bi = emergent(pain_only, deficits)
    ranked = sorted(((t, g) for t, g in groups.items() if g["reviews"]),
                    key=lambda x: len(x[1]["reviews"]), reverse=True)

    out_json = {"total": len(reviews), "filter": filt, "pain_only": len(pain_only), "deficits": {}}
    for t, g in ranked:
        # 상황 부합 원문 최대 N개 — 감정 강한/저평점 순, 중복 문구 제거(원문 자체가 광고 소재)
        seen_q, uniq = set(), []
        for q in sorted(g["quotes"], key=lambda q: (-q["emotion"], q.get("rating", 0))):
            key = (q.get("deficit") or "").strip()
            if key and key not in seen_q:
                seen_q.add(key); uniq.append(q)
        out_json["deficits"][t] = {
            "count": len(g["reviews"]), "kind": g["quotes"][0]["kind"],
            "avg_rating": round(sum(_rating(r) for r in g["reviews"]) / len(g["reviews"]), 2),
            "quote_n": len(uniq[:args.quotes]),
            "top_quotes": uniq[:args.quotes],
        }
    out_json["emergent_unigrams"], out_json["emergent_bigrams"] = uni, bi
    json.dump(out_json, open(args.out + ".json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)

    D = out_json["deficits"]
    L = [f"# 위닝 소재 추출 — 결핍-우선 · 분석 {filt['analyzed']}건 (진짜 고객 VOC)", "",
         f"> 필터: 전체 {filt['raw']}건 → 별점-only(본문없음) {filt['empty_body']} 제외 · "
         f"체험단·협찬 {filt['promo']} 제외 → **분석 {filt['analyzed']}건**.",
         "> 관점: 제품 강점이 아니라 **리뷰어의 공통 문제(결핍·컴플렉스)**를 상황별로 묶어 건수로 센다.",
         "> **공통 문제일수록(건수↑) 광고 소구·후킹 우선순위가 높다.**",
         "> TRAIT=고객이 원래 가진 조건(직모·짧음·힘없음 등, 위닝 핵심) · OUTCOME=증상(번짐 등, 부정어 제외 후만).", ""]

    L += ["## ① 리뷰어의 문제(결핍) — 빈도순 = 소구 강도순", "",
          "| 결핍 상황 | 유형 | 출현 | 평균★ | 대표 원문(가장 절박한) |", "|---|:--:|--:|--:|---|"]
    for t, d in D.items():
        q = d["top_quotes"][0]["deficit"] if d["top_quotes"] else ""
        q = (q[:48] + "…") if len(q) > 48 else q
        L.append(f"| {t} | {d['kind']} | {d['count']} | {d['avg_rating']} | {q} |")
    L.append("")

    L += [f"## ② 공통 문제별 — 상황 부합 원문 (최대 {args.quotes}개, 광고 소재용)", "",
          "> 원문 자체가 카피 재료다. 상황에 부합하는 고객 원문을 최대한 많이 실어 VOC 반영도를 높인다.", ""]
    for t, d in list(D.items())[:args.top]:
        L.append(f"### {t} — 결핍 {d['count']}건 · 부합 원문 {d['quote_n']}개 ({d['kind']})")
        for q in d["top_quotes"]:
            if q["deficit"]:
                star = f"{q.get('rating','')}★"
                L.append(f"- ({star}) “{q['deficit'].strip()}”")
        L.append("")

    L += ["## ③ 위닝 소재 — 소구포인트 · 후킹 재료 (반복 높은 순)", "",
          "> 후킹 재료 = 고객이 자기 입으로 말한 결핍 원문. LLM/카피라이터가 이걸 광고 훅으로 다듬는다.", ""]
    for i, (t, d) in enumerate(list(D.items())[:args.top], 1):
        seed = next((q for q in d["top_quotes"] if q["deficit"]), None)
        raw = f'“{seed["deficit"].strip()}”' if seed else "-"
        L += [f"{i}. **{t}** — 결핍 {d['count']}건 · 평균 {d['avg_rating']}★",
              f"   - 타깃(이런 분): {t} 겪는 사람",
              f"   - 후킹 재료(고객 원문): {raw}",
              f"   - 소구 각도: 이 결핍이 {d['count']}회 반복 → 상세페이지/광고 히어로에 '{t}' 해결을 전면에"]
    L.append("")

    if uni or bi:
        L += ["## ④ 발견된 결핍(사전 밖 · 창발) — 새 소구축 후보",
              "- 단어: " + ", ".join(f"{w}({n})" for w, n in uni[:15] if n > 1),
              "- 연어: " + ", ".join(f"{w}({n})" for w, n in bi[:12] if n > 1), ""]
    L.append(f"> 통증 마커만 있고 사전 미분류: {len(pain_only)}건 — 창발 표현을 --deficits 사전에 넣으면 흡수.")

    open(args.out + ".md", "w", encoding="utf-8").write("\n".join(L))
    if args.xlsx:
        write_xlsx(args.out + ".xlsx", groups, ranked, D)

    print(json.dumps({"filter": filt,
                      "deficits_ranked": [[t, d["count"], d["kind"]] for t, d in D.items()],
                      "pain_only": len(pain_only),
                      "out_md": args.out + ".md", "out_json": args.out + ".json"}, ensure_ascii=False))


def write_xlsx(path, groups, ranked, D):
    import openpyxl
    from openpyxl.styles import Font
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "결핍 요약"
    ws.append(["결핍 상황", "유형", "출현", "평균★", "대표 결핍 원문"])
    for c in ws[1]:
        c.font = Font(bold=True)
    for t, d in D.items():
        seed = next((q for q in d["top_quotes"] if q["deficit"]), None)
        ws.append([t, d["kind"], d["count"], d["avg_rating"], seed["deficit"] if seed else ""])
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["E"].width = 70
    cols = ["rating", "user", "결핍 원문", "해결 원문", "감정강도"]
    for t, g in ranked:
        s = wb.create_sheet(_safe(t))
        s.append(cols)
        for c in s[1]:
            c.font = Font(bold=True)
        s.freeze_panes = "A2"
        for q in sorted(g["quotes"], key=lambda q: q["emotion"], reverse=True):
            s.append([q["rating"], q["user"], q["deficit"], q["resolve"], q["emotion"]])
        s.column_dimensions["C"].width = 55
        s.column_dimensions["D"].width = 55
    wb.save(path)


def _safe(name):
    for ch in "[]:*?/\\":
        name = name.replace(ch, " ")
    return name[:31]


if __name__ == "__main__":
    main()
